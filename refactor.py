import pandas as pd
import threading
import time
import os
from tqdm import tqdm
from rich import print
from rich.progress import Progress
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, NamedStyle


surgery = "Cirugia"
outpatient_clinic = "Consulta externa"
internment = "Internacion"
urgencies = "Urgencias"
raw_cid10 = "raw_CID10"
cid10 = "CID10"
cid10_file = cid10

def load_data(file_name):
    """Load data from an Excel file based on the given file name."""
    file_path = None

    if file_name == surgery:
        file_path = "Data-cruda/BD Cirugía.xlsx"
    elif file_name == outpatient_clinic:
        file_path = "Data-cruda/BD Consulta externa.xlsx"
    elif file_name == internment:
        file_path = "Data-cruda/BD Internación.xlsx"
    elif file_name == urgencies:
        file_path = "Data-cruda/BD Urgencias.xlsx"
    elif file_name == raw_cid10:
        file_path = "CIE10\CIE10 AGRUPADO.xlsx"
    elif file_name == cid10:
        file_path = "CIE10\CID10_clean.xlsx"

    df = pd.read_excel(file_path)
    return df

def create_CID10():
    df = load_data(raw_cid10)
    df = delete_columns(raw_cid10, df)
    df = df.rename(columns={'COD_4.1': 'Codigo_Diagnostico'})
    df.to_excel('CIE10/' + str(cid10) + '_clean.xlsx', index=False)
    # print(df.head()) 
    return df

def delete_columns(file_name, df):
    """
    Delete columns from the given DataFrame based on the given file name.
    """
    # Define the columns to keep
    keep_columns = {
        surgery: ['VALIDAR REPETIDOS', 'Codigo DX Principal', 'Diagnostico Principal'],
        outpatient_clinic: ['VALIDACIÓN JUNTA DIRECTIVA Y PRODUCCIÓN AMBULATORIO', 'Cod Diag Princl', 'Diag.Princ'],
        internment: ['VALIDACION INTERNACION', 'Código diagnóstico alta', 'Texto diagnóstico alta'],
        urgencies: ['Clase de consulta', 'Cod Diag Princl', 'Diag.Princ'],
        raw_cid10: ['COD_4.1', 'CATEGORÍA', 'SUBCATEGORÍA']
    }[file_name]

    # Filter the columns
    columns = [col for col in keep_columns if col in df.columns]
    if columns:
        df = df[columns]
    else:
        print("No se encontraron las columnas especificadas")

    return df

def remove_rows(file_name, df):
    if(file_name == surgery):
        df = df[df['VALIDAR REPETIDOS'] == 'NO REPETIDO']
    elif(file_name == outpatient_clinic):
        df = df[df['VALIDACIÓN JUNTA DIRECTIVA Y PRODUCCIÓN AMBULATORIO'] == 'TENER EN CUENTA J. DIRECTIVA']
    elif(file_name == internment):
        df = df[df['VALIDACION INTERNACION'] == 'INTERNACION']
    elif(file_name == urgencies):
        df = df[df['Clase de consulta'] == 'Urgencias']
    # print(df.head()) 
    return df

def count_codes(file_name, df):
    column_mappings = {
        surgery: ('Codigo DX Principal', 'Diagnostico Principal'),
        outpatient_clinic: ('Cod Diag Princl', 'Diag.Princ'),
        internment: ('Código diagnóstico alta', 'Texto diagnóstico alta'),
        urgencies: ('Cod Diag Princl', 'Diag.Princ')
    }

    if file_name in column_mappings:
        code_col, diag_col = column_mappings[file_name]

        # Group by code column and count the occurrences
        df_counted = df.groupby(code_col)[diag_col].count().reset_index()
        
        # Rename the count column
        df_counted = df_counted.rename(columns={diag_col: 'Count'})

        # Add the original diagnosis column if it exists
        if diag_col in df.columns:
            df_counted = df_counted.merge(df[[code_col, diag_col]].drop_duplicates(), on=code_col)

        # Sort results by count in descending order
        df_sorted = df_counted.sort_values(by='Count', ascending=False)

        # Rename the code column al final
        df_sorted = df_sorted.rename(columns={code_col: 'Codigo_Diagnostico'})
        df_sorted = df_sorted.rename(columns={diag_col: 'Diagnostico'})

        # Save to Excel file
        # df_sorted.to_excel(str(file_name)+'.xlsx', index=False)
        
        return df_sorted

def add_categories(df1, df2):
    # Elimina las filas duplicadas en la columna Codigo_Diagnostico
    df2 = df2.drop_duplicates(subset='Codigo_Diagnostico')

    # Crea un diccionario que mapea los códigos de diagnóstico con las categorías y subcategorías
    mapping = df2.set_index('Codigo_Diagnostico')[['CATEGORÍA', 'SUBCATEGORÍA']].to_dict(orient='index')

    # Utiliza el método map para buscar los valores de las columnas CATEGORÍA y SUBCATEGORÍA en df2 y copiarlos en df1
    df1['CATEGORÍA'] = df1['Codigo_Diagnostico'].map(lambda x: mapping.get(x, {}).get('CATEGORÍA'))
    df1['SUBCATEGORÍA'] = df1['Codigo_Diagnostico'].map(lambda x: mapping.get(x, {}).get('SUBCATEGORÍA'))

    # print(df1.head()) 
    # df1.to_excel('output.xlsx', index=False)
    return df1

def order_by_category(df):
    # Calcula la sumatoria de la columna "COUNT" para cada grupo de categorías
    df_grouped = df.groupby('CATEGORÍA')['Count'].sum().reset_index()
    df_grouped.columns = ['CATEGORÍA', 'SUM_COUNT']

    # Ordena los grupos por la sumatoria en orden descendente
    df_grouped = df_grouped.sort_values(by='SUM_COUNT', ascending=False)

    # Une el DataFrame original con el DataFrame de grupos ordenados
    df_ordered = pd.merge(df, df_grouped, on='CATEGORÍA')

    # Ordena el DataFrame resultante por la sumatoria y luego por la columna "COUNT" en orden descendente
    df_ordered = df_ordered.sort_values(by=['SUM_COUNT', 'Count'], ascending=[False, False])
    
    # df_ordered.to_excel('output.xlsx', index=False)
    return df_ordered

def add_percentage_column(df):
    df['PERCENTAGE_GROUP'] = (df['Count'] / df['SUM_COUNT']) * 100
    # df.to_excel('output.xlsx', index=False)
    return df

def add_total_count(df):
    total_count = df['Count'].sum()
    df['TOTAL_COUNT'] = total_count
    # df.to_excel('output.xlsx', index=False)
    return df

def add_group_percentage_column(file_name, df):
    total_count = df['TOTAL_COUNT'].iloc[0]
    df['GROUP_PERCENTAGE'] = (df['SUM_COUNT'] / total_count) * 100
    df.to_excel('Data-procesada/' + str(file_name) + '_processed.xlsx', index=False)
    return df

def show_instructions():
    print("\n[!] Instrucciones:")
    print("\n[!] Nombre de archivo CID10 base:")
    print("     CIE10\CIE10 AGRUPADO.xlsx\n")
    print("[!] Nombres de archivos base:")
    print("     Data-cruda\BD Cirugía.xlsx")
    print("     Data-cruda\BD Consulta externa.xlsx.xlsx")
    print("     Data-cruda\BD Internación.xlsx.xlsx")
    print("     Data-cruda\BD Urgencias.xlsx\n")

def mostrar_menu():
    print("[?] ¿Qué deseas hacer?")
    print("[1] Crear archivo CID10")
    print("[2] Cargar archivo CID10")
    print("[3] Salir\n")

def mostrar_menu_servicios():
    print("\n[?] ¿Qué deseas hacer?")
    print("[1] Procesar Cirugía")
    print("[2] Procesar Consulta Externa")
    print("[3] Procesar Internación")
    print("[4] Procesar Urgencias")
    print("[5] Salir\n")

class Imprimir:
    def __init__(self):
        self.lock = threading.Lock()

    def imprimir(self, mensaje, end='\n'):
        with self.lock:
            print(mensaje, end=end)

def mostrar_progreso(event, imprimir):
    spinner_chars = ['|', '/', '-', '\\']
    i = 0
    while not event.is_set():
        char = spinner_chars[i % len(spinner_chars)]
        imprimir.imprimir(f"[!] Procesando...{char}", end='\r')
        i += 1
        time.sleep(0.1)

def process_service(file_name, df2):
    imprimir = Imprimir()

    # Crear un evento para controlar el hilo del spinner
    event = threading.Event()

    # Crear un hilo para el spinner
    hilo_progreso = threading.Thread(target=mostrar_progreso, args=(event, imprimir))
    hilo_progreso.daemon = True
    hilo_progreso.start()

    # Ejecutar tareas (simuladas con sleep para dar tiempo a ver la animación)
    df = load_data(file_name)
    df = delete_columns(file_name, df)
    df = remove_rows(file_name, df)
    df = count_codes(file_name, df)
    df = add_categories(df, df2)
    df = order_by_category(df)
    df = add_percentage_column(df)
    df = add_total_count(df)
    df = add_group_percentage_column(file_name, df)

    df = combine_cells(df, file_name)
    # Cuando se termina el procesamiento, señalizar al spinner que se detenga
    event.set()

def ensure_directories_exist():
    """
    Verifica si existen las carpetas 'CIE10', 'Data-cruda', y 'Data-procesada'.
    Si no existen, las crea.
    """
    directories = ["CIE10", "Data-cruda", "Data-procesada"]

    for directory in directories:
        if not os.path.exists(directory):
            os.makedirs(directory)
            print(f"Directorio '{directory}' creado exitosamente.")
        else:
            print(f"Directorio '{directory}' ya existe.")



def combine_cells(df, file_name):
    """
    Combina todas las celdas de la columna 'TOTAL_COUNT', luego combina las celdas de 'SUM_COUNT', 'GROUP_PERCENTAGE',
    y 'CATEGORÍA' solo cuando tienen la misma 'CATEGORÍA'.
    Retorna un DataFrame actualizado.
    :param df: DataFrame con los datos.
    :return: DataFrame actualizado con las celdas combinadas reflejadas.
    """
    if not isinstance(df, pd.DataFrame):
        raise ValueError("El argumento 'df' debe ser un DataFrame.")

    # Crear un archivo de Excel desde el DataFrame
    wb = Workbook()
    ws = wb.active

    # Insertar encabezados y datos del DataFrame en el archivo de Excel
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=1):
        ws.append(row)

    # Obtener los índices de las columnas 'TOTAL_COUNT', 'SUM_COUNT', 'GROUP_PERCENTAGE' y 'CATEGORÍA'
    header_row = 1
    total_count_col = None
    sum_count_col = None
    group_percentage_col = None
    category_col = None

    for col_idx, cell in enumerate(ws[header_row], start=1):
        if cell.value == "TOTAL_COUNT":
            total_count_col = col_idx
        if cell.value == "SUM_COUNT":
            sum_count_col = col_idx
        if cell.value == "GROUP_PERCENTAGE":
            group_percentage_col = col_idx
        if cell.value == "CATEGORÍA":
            category_col = col_idx

    if not total_count_col or not sum_count_col or not group_percentage_col or not category_col:
        raise ValueError("No se encontraron las columnas 'TOTAL_COUNT', 'SUM_COUNT', 'GROUP_PERCENTAGE' o 'CATEGORÍA' en el DataFrame.")

    # Combinar todas las celdas en la columna 'TOTAL_COUNT'
    start_row = header_row + 1
    ws.merge_cells(start_row=start_row, start_column=total_count_col, 
                   end_row=ws.max_row, end_column=total_count_col)

    # Centrar el texto de la celda fusionada en 'TOTAL_COUNT'
    merged_cell = ws.cell(row=start_row, column=total_count_col)
    merged_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Combinar celdas en 'SUM_COUNT', 'GROUP_PERCENTAGE', y 'CATEGORÍA' solo si tienen la misma 'CATEGORÍA'
    current_category = None
    start_merge_row = start_row

    for row in range(start_row, ws.max_row + 1):
        category_value = ws.cell(row=row, column=category_col).value

        if category_value != current_category:
            # Fusionar celdas anteriores si corresponde
            if current_category is not None and row - 1 > start_merge_row:
                for col in [sum_count_col, group_percentage_col, category_col]:
                    ws.merge_cells(start_row=start_merge_row, start_column=col, 
                                   end_row=row - 1, end_column=col)
                    # Centrar el texto de la celda fusionada
                    merged_cell = ws.cell(row=start_merge_row, column=col)
                    merged_cell.alignment = Alignment(horizontal="center", vertical="center")

            # Actualizar la categoría actual y la fila de inicio de fusión
            current_category = category_value
            start_merge_row = row

    # Fusionar las últimas celdas si corresponde
    if current_category is not None and ws.max_row > start_merge_row:
        for col in [sum_count_col, group_percentage_col, category_col]:
            ws.merge_cells(start_row=start_merge_row, start_column=col, 
                           end_row=ws.max_row, end_column=col)
            merged_cell = ws.cell(row=start_merge_row, column=col)
            merged_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Convertir de vuelta a DataFrame para devolver
    data = ws.values
    columns = next(data)  # Tomar encabezados
    df_updated = pd.DataFrame(data, columns=columns)

    # Guardar el archivo Excel
    output_path = "Data-procesada/" + str(file_name) + "_processed.xlsx"
    wb.save(output_path)
    print(f"Archivo con formato de porcentaje guardado en {output_path}")

    return df_updated


def main():
    ensure_directories_exist()
    show_instructions()
    show_first_menu = True
    while True:
        if show_first_menu:
            mostrar_menu()
            show_first_menu = False
            opcion = input("[!] Elija una opción: ")
            if opcion == "1":
                # Lógica para crear el archivo CID10
                print("[!] Creando archivo CID10...")
                create_CID10()
                df2 = load_data(cid10_file)
                print("\n[!] Archivo CID10 creado exitosamente.")
                print("[!] Archivo CID10 cargado exitosamente.\n")
            elif opcion == "2":
                # Lógica para cargar el archivo CID10
                print("[!] Cargando archivo CID10...")
                df2 = load_data(cid10_file)
                print("[!] Archivo CID10 cargado exitosamente.\n")
            elif opcion == "3":
                print("[!] Saliendo...\n")
                break
            else:
                print("[!] Opción inválida. Por favor, elija una opción válida.\n")
                show_first_menu = True
        if show_first_menu is False:
            mostrar_menu_servicios()
            opcion_servicio = input("[!] Elija una opción: ")
            if opcion_servicio == "1":
                file_name = surgery
            elif opcion_servicio == "2":
                file_name = outpatient_clinic
            elif opcion_servicio == "3":
                file_name = internment
            elif opcion_servicio == "4":
                file_name = urgencies
            elif opcion_servicio == "5":
                print("[!] Saliendo...\n")
                break
            else:
                print("[!] Opción inválida. Por favor, elija una opción válida.\n")
                continue
            print(f"[!] Procesando servicio: {file_name} ")
            process_service(file_name, df2)
            print("[!] Procesamiento realizado exitosamente.\n")

if __name__ == "__main__":
    main()

